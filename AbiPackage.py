import csv
import os.path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk

class Data: # Every line of the file will be made into data class which we use later

    def __init__(self, code: str, postision: str, version: str = None) -> None:
        self.code = code
        self.version = version
        self.positsion = postision

    def getCode(self) -> str:
        return self.code
    
    
    def getVersion(self):
        return self.version
    
    def getPosition(self) -> str:
        return self.positsion
    
    
class DataBase: # This where we will hold out data


    def __init__(self) -> None:
        self.info = []

    
    def add(self, item: Data):
        self.info.append(item)

    def getDataBase(self):
        return self.info
    
    def __add__(self, other):
        if isinstance(other, DataBase):
            self.info.extend(other.getDataBase())
        return self
    
    def getPositions(self) -> list: # This method returns all positsion it has in it its database used for later in magic command
        result = []
        for item in self.info:
            if item.getPosition() not in result:
                result.append(item.getPosition())
        return result


    # When working with CSV files
    def getData(self, NameOfTheFIle: str) -> None:
        # Opens the file.
        with open(NameOfTheFIle, encoding='utf-8-sig',) as csv_file:

            # Saves the file as a csv.reader object and separates the lines in file to lists of strings which were separated by the delimiter.
            csv_reader = list(csv.reader(csv_file, delimiter=";"))
            check = self.DatabaseFiller(csv_reader)
            if check == -1:
                return -1
            else:
                return 1
    # When working for XLSX files
    def getDataWithXlsx(self, NameOfTheFile: str) -> int:

        workbook = load_workbook(NameOfTheFile)
        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        check = self.DatabaseFiller(data)
        if check == -1:
            return -1
        else:
            return 1



    def DatabaseFiller(self, items: list) -> int:
        startingIndex = self.tableStartingPoint(items)
        NoneCount = 0
        if startingIndex == -1:
            return -1
        if "Ver" in items[startingIndex]:
            versionIndex =  items[startingIndex].index("Ver")
        else:
            versionIndex =  None
        koodIndex = items[startingIndex].index("Code")
        positsioonIndex = items[startingIndex].index("Ref")
        for row in items[startingIndex+1:]:
            if row[koodIndex] == None or row[koodIndex] == "":
                continue
            if row[positsioonIndex] == None:
                item = Data(row[koodIndex], f"None{NoneCount}", row[versionIndex])
                NoneCount += 1
                self.add(item)
            else:
                positsions = row[positsioonIndex].split(",")
                for positsion in positsions:
                    if versionIndex == None:
                        item = Data(row[koodIndex].strip(), positsion.strip())    
                    else:
                        item = Data(row[koodIndex].strip(), positsion.strip(), row[versionIndex])
                    self.add(item)
        return 1


    def tableStartingPoint(self, array: list) -> int:
        for i in range(len(array)):
            if "Code" in array[i]:
                return i
        return -1

def magic(file_first: str, file_second: str, output: str) -> int:


    if not os.path.isfile(file_first):
        return -1
    if not os.path.isfile(file_second):
        return 0
    

    workbook = Workbook()
    sheet = workbook.active


    red = Font(color="FF0000")
    black = Font(color="000000")
    green = PatternFill(start_color="5bba75", end_color="5bba75", fill_type="solid")
    purple = PatternFill(start_color="c95da0", end_color="c95da0", fill_type="solid")
    
    if file_first.split(".")[-1] == "csv":
        file1 = DataBase()
        check = file1.getData(file_first)
        if check == -1:
            return 404
        fileUnchanged1 = file1.getDataBase()


        file2 = DataBase()
        check = file2.getData(file_second)
        if check == -1:
            return 404
        fileUnchanged2 = file2.getDataBase()
    if file_first.split(".")[-1] == "xlsx":
        file1 = DataBase()
        check = file1.getDataWithXlsx(file_first)
        if check == -1:
            return 404
        fileUnchanged1 = file1.getDataBase()


        file2 = DataBase()
        check = file2.getDataWithXlsx(file_second)
        if check == -1:
            return 404
        fileUnchanged2 = file2.getDataBase()

    positions = list(set((file1.getPositions() + file2.getPositions())))
    positions.sort()
    rows = [["Code", "Version", "Code", "Version", "Ref"]]


    for position in positions:
        foundFile1 = list(filter(lambda x: x.getPosition() == position, fileUnchanged1))
        foundFile2 = list(filter(lambda x: x.getPosition() == position, fileUnchanged2))
        if len(foundFile1) > 0 and len(foundFile2) > 0:
            row = [foundFile1[0].getCode(), foundFile1[0].getVersion()
                    , foundFile2[0].getCode(), foundFile2[0].getVersion(), foundFile2[0].getPosition()]
        if len(foundFile1) > 0 and len(foundFile2) == 0:
            row = [foundFile1[0].getCode(), foundFile1[0].getVersion(), "", "", foundFile1[0].getPosition()]
        if len(foundFile1) == 0 and len(foundFile2) > 0:
            row = ["", "", foundFile2[0].getCode(), foundFile2[0].getVersion(), foundFile2[0].getPosition()]
        rows.append(row)


    header_row = rows[0]
    for col_num, value in enumerate(header_row, start=1):
        cell = sheet.cell(row=1, column=col_num)
        if col_num == 1 or col_num == 2:
            cell.fill = green
        if col_num == 3 or col_num == 4:
            cell.fill = purple
        cell.value = value

    data = rows[1:]
    for row_num, row in enumerate(data, start=2):
        for col_num, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num in [1, 3]:  # Columns 1 and 3
                if col_num == 1:
                    cell.fill = green
                    if(row[col_num - 1] != row[col_num + 1]):
                        cell.font = red
                if col_num == 3:
                    cell.fill = purple
                    if(row[col_num - 1] != row[col_num - 3]):
                        cell.font = red
            elif col_num in [2, 4]:  # Columns 2 and 4
                if col_num == 2:
                    cell.fill = green
                    if(row[col_num - 1] != row[col_num + 1]):
                        cell.font = red
                if col_num == 4:
                    cell.fill = purple
                    if(row[col_num - 1] != row[col_num - 3]):
                        cell.font = red
            elif col_num == 5:  # Column 5
                cell.font = black
    
    workbook.save(output)
    return 1
###################################################
'''
This is where the GUI implementation begins

'''
###################################################
def on_entry_click(event):
    if entry1.get() == "Enter file name":
        entry1.delete(0, tk.END)


def on_entry_leave(event):
    if entry1.get() == "":
        entry1.insert(0, "Enter file name")


def on_entry2_click(event):
    if entry2.get() == "Enter file name":
        entry2.delete(0, tk.END)


def on_entry2_leave(event):
    if entry2.get() == "":
        entry2.insert(0, "Enter file name")


def on_entry3_click(event):
    if entry3.get() == "Enter output file name":
        entry3.delete(0, tk.END)


def on_entry3_leave(event):
    if entry3.get() == "":
        entry3.insert(0, "Enter output file name")

def CheckFileTypes() -> bool:
    return entry1.get().split(".")[-1] == entry2.get().split(".")[-1]

def CheckFileReadble() -> bool:
    types = ["csv", "xlsx"]
    return entry1.get().split(".")[-1] in types and entry2.get().split(".")[-1] in types



def button_command(): # This is where all the work is done when button is pressed
    def close_popup():
        if popup.winfo_exists():  # Check if the popup exists before destroying it
            popup.destroy()

    def close_popup_and_main_window(): # Since otherwise the code didn't work I had to put them into this function.
        close_popup()
        window.destroy()

    popup = tk.Toplevel()

    if not CheckFileTypes(): # Checks if both files are of same type
        popup.title("Type Error")

        label = tk.Label(popup, text="PLease make sure that boths files are of same types")
        label.pack(padx=10, pady=10)

        close_button = tk.Butoon(popup, text="Close", command=close_popup)

    if not CheckFileReadble: # Checks if both files are csv or xlsx types
        popup.title("Type Error")

        label = tk.Label(popup, text="Files have to be either csv or xlsx")
        label.pack(padx=10, pady=10)

        close_button = tk.Butoon(popup, text="Close", command=close_popup)
    
    result = magic(entry1.get(), entry2.get(), entry3.get())

    if result == 404:
        popup.title("Error 404")

        label = tk.label(popup, text="Inside one of the files the table was not found")
        label.pack(padx=10, pady=10)

        close_button = tk.Button(popup, text="Close", command=close_popup)
        close_button.pack(pady=10)

    if result == 1: # Magic function returns a number indicating how far did it make towards the end result with 1 indicating it fully completed the job
        popup.title("Success")

        label = tk.Label(popup, text=f"Operation was completed successfully and the result was saved to {entry3.get()}")
        label.pack(padx=10, pady=10)

        close_button = tk.Button(popup, text="Close", command=close_popup_and_main_window)
        close_button.pack(pady=10)

    elif result == -1: # -1 indicates that it did not find the first file that was entered into magic function.
        popup.title("Error")

        label = tk.Label(popup, text=f"File {entry1.get()} not found!")
        label.pack(padx=10, pady=10)

        close_button = tk.Button(popup, text="Close", command=close_popup)
        close_button.pack(pady=10)

    elif result == 0: # 0 indicates that it did not find the second file that was entered into magic function.
        popup.title("Error")

        label = tk.Label(popup, text=f"File {entry2.get()} not found")
        label.pack(padx=10, pady=10)

        close_button = tk.Button(popup, text="Close", command=close_popup)
        close_button.pack(pady=10)

    popup.deiconify()



if __name__ == "__main__":
    # Create the main window
 # Create the main window
    window = tk.Tk()
    window.title("Difference finder v1.6")
    window.geometry("300x200")

    # Create the first text field (Entry widget) with default text
    entry1 = tk.Entry(window)
    entry1.insert(0, "Enter file name")  # Set default text

    # Bind the entry field events
    entry1.bind("<FocusIn>", on_entry_click)
    entry1.bind("<FocusOut>", on_entry_leave)

    # Create the second text field (Entry widget) with default text
    entry2 = tk.Entry(window)
    entry2.insert(0, "Enter file name")  # Set default text

    entry2.bind("<FocusIn>", on_entry2_click)
    entry2.bind("<FocusOut>", on_entry2_leave)

    
    # Create the third text field with default text
    entry3 = tk.Entry(window)
    entry3.insert(0, "Enter output file name") # Set default text

    entry3.bind("<FocusIn>", on_entry3_click)
    entry3.bind("<FocusOut>", on_entry3_leave)

    
    button = tk.Button(window, text="Show differences", command=button_command)

    entry1.pack()
    entry2.pack()
    entry3.pack()
    button.pack()

    window.mainloop()
    
    
            





